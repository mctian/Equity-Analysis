from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt


def getValidTickerIndexes(worksheet):
    columns = []
    row = worksheet['{}{}:{}{}'.format('A', 3, 'IYS', 3)]
    for n, cell in enumerate(row[0]):
        if n % 2 == 1 and cell.value is not None:
            columns.append(n)
    return columns


# unlike the other functions, this one should use worksheet "stock list" rather than the second ws
def findTickerIndex(ticker, worksheet):
    for n, row in enumerate(worksheet.iter_rows()):
        for cell in row:
            if str(cell.value) == ticker:
                return n * 2 + 1
    print(ticker + " not found.")


# Excel indexes start with 1 rather than 0
def convertIndexToLetter(index):
    i = index
    letterIndex = ""

    while i > 0:
        remainder = i % 26
        if remainder == 0:
            letterIndex += 'Z'
            i = i/26 - 1
        else:
            letterIndex += (chr((remainder-1) + ord('A')))
            i = i/26

    return ''.join(reversed(letterIndex))


def getAllTimeSeries(worksheet):
    indexes = getValidTickerIndexes(worksheet)
    for i in indexes:
        dates = []
        values = []
        row_range = 100
        letterIndex = convertIndexToLetter(i)
        for cell in worksheet['{}{}:{}{}'.format(letterIndex, 4, letterIndex, row_range)]:
            if cell[0].value is not None:
                dates.append(cell[0].value)
            else:
                break
        letterIndex = convertIndexToLetter(i+1)
        for cell in worksheet['{}{}:{}{}'.format(letterIndex, 4, letterIndex, row_range)]:
            if cell[0].value is not None:
                values.append(cell[0].value)
            else:
                break
        dates = np.array(dates, dtype = np.datetime64)
        data = pd.Series(values, index = dates)
    return


def getTimeSeries(stock, worksheet, factor):
    stockList = wb["stock list"]
    i = findTickerIndex(stock, stockList)
    dates = []
    values = []
    row_range = 400
    letterIndex = convertIndexToLetter(i)
    for cell in worksheet['{}{}:{}{}'.format(letterIndex, 4, letterIndex, row_range)]:
        if cell[0].value is not None:
            dates.append(cell[0].value)
        else:
            break
    letterIndex = convertIndexToLetter(i+1)
    for cell in worksheet['{}{}:{}{}'.format(letterIndex, 4, letterIndex, row_range)]:
        if cell[0].value is not None:
            values.append(cell[0].value)
        else:
            break
    dates = np.array(dates, dtype = np.datetime64)
    data = pd.Series(values, index = dates, name = factor)
    return data


if __name__ == "__main__":
    wb = load_workbook(filename = "USEquity(Price to Book Ratio).xlsm")
    ws = wb["Price to Book"]
    pbr = getTimeSeries("AAPL", ws, "PBR")
    wb = load_workbook(filename = "USEquity(Dividend Yield).xlsm")
    ws = wb["Dividend Yield"]
    dy = getTimeSeries("AAPL", ws, "Div Yield")
    plt.show()
