from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import time #clock, time methods


# returns a list of valid indexes in integers, starting from 1
def getValidTickerIndexes(worksheet):
    columns = []
    row = worksheet['{}{}:{}{}'.format('A', 3, 'IYS', 3)]
    for n, cell in enumerate(row[0]):
        if n % 2 == 1 and cell.value is not None:
            columns.append(n)
    return columns


# unlike the other functions, this one should use worksheet "stock list" rather than the second ws
def findTickerIndex(ticker, worksheet):
    for n, row in enumerate(worksheet.iter_rows()):
        for cell in row:
            if str(cell.value) == ticker:
                return n * 2 + 1
    print(ticker + " not found.")
    return


# Excel indexes start with 1 rather than 0
def convertIndexToLetter(index):
    i = index
    letterIndex = ""
    while i > 0:
        remainder = (i - 1) % 26
        letterIndex = (chr((remainder) + ord('A'))) + letterIndex
        i = int((i - remainder)/26)
    return letterIndex


# returns list of all tickers
def getAllTickers(worksheet):
    tickers = []
    row = worksheet['{}{}:{}{}'.format('A', 3, 'IYS', 3)]
    for n, cell in enumerate(row[0]):
        if n % 2 == 1 and cell.value is not None:
            tickers.append(str(worksheet[convertIndexToLetter(n)+"1"].value))
    df = pd.DataFrame(tickers, columns=['Tickers'])
    df.to_csv("stocklist.csv", index=False)
    return tickers


def getAllTickersBySector(worksheet):
    sectors = ['Consumer Discretionary', 'Consumer Staples', 'Energy', 'Financials', 'Health Care', 'Industrials', 'Information Technology', 'Materials', 'Real Estate', 'Telecommunication Services', 'Utilities']
    df = pd.read_csv("stocklist.csv",index_col = 0)
    stocks = set(df.index.tolist())
    print(stocks)
    for sector in sectors:
        tickers = []
        row = worksheet['{}{}:{}{}'.format('A', 3, 'IYS', 3)]
        for n, cell in enumerate(row[0]):
            print(str(worksheet[convertIndexToLetter(n+1)+"1"].value))
            if n % 2 == 0 and cell.value == sector and str(worksheet[convertIndexToLetter(n+1)+"1"].value) in stocks:
                print(cell.value)
                tickers.append(str(worksheet[convertIndexToLetter(n+1)+"1"].value))
        df = pd.DataFrame(tickers, columns=['Tickers'])
        df.to_csv(sector+".csv", index=False)


# merges date indexes of a csv with deltas < days
def mergeCloseIndexes(tickers, days):
    for ticker in tickers:
        df = pd.read_csv(ticker+".csv", index_col = 0)
        df.index = pd.to_datetime(df.index)
        indexToDelete = []
        for i in range(1,len(df.index)):
            if abs(df.index[i] - df.index[i-1]) < datetime.timedelta(days):
                row1 = df.iloc[i-1]
                row2 = df.iloc[i]
                rowUpdated = row1.combine_first(row2)
                df.iloc[i-1] = rowUpdated
                indexToDelete = [i] + indexToDelete
        df.drop(df.index[indexToDelete], inplace=True)
        f = open(ticker+".csv", "w+")
        f.close()
        df.to_csv(ticker+".csv")
    return


# outputs all time series to csv files
def getAllTimeSeries(worksheet, factor):
    indexes = getValidTickerIndexes(worksheet)
    for i in indexes:
        name = str(worksheet[convertIndexToLetter(i)+"1"].value)
        dates = []
        values = []
        row_range = 800
        letterIndex = convertIndexToLetter(i)
        for cell in worksheet['{}{}:{}{}'.format(letterIndex, 4, letterIndex, row_range)]:
            if cell[0].value is not None:
                dates.append(cell[0].value)
            else:
                break
        letterIndex = convertIndexToLetter(i+1)
        for cell in worksheet['{}{}:{}{}'.format(letterIndex, 4, letterIndex, row_range)]:
            if cell[0].value is not None:
                values.append(cell[0].value)
            else:
                break
        dates = np.array(dates, dtype = np.datetime64)
        try:
            data = pd.Series(values, index = dates)
            outputTimeSeries(data, factor, name)
        except ValueError:
            print(factor)
            print("Value Error for: " + name)
            print("# dates: " + str(len(dates)))
            print("# vals: " + str(len(values)))
    return


# outputs an inputed time series to csv file
def outputTimeSeries(series, factor, name):
    dfSeries = pd.DataFrame({factor: series})
    try:
        df = pd.read_csv(name+".csv", index_col = 0)
        df.index = pd.to_datetime(df.index)
    except FileNotFoundError:
        df = pd.DataFrame()
    df = pd.concat([df, dfSeries], axis = 1)
    df.to_csv(name+".csv")
    return


# outputs all time series for all factors
def outputAllFactors():
    factors = ["EPS Growth", "Trailing EPS", "Total Debt to Total Equity", "Volatility 180 Day", "Return on Invested Capital", "Return on Common Equity", "Return on Assets", "Volatility 360 Day", "Earnings Momentum", "Price to Cash Flow", "Price to Book Ratio", "EPS", "Dividend Yield", "Forward PE", "Volume", "Last Price"]
    filenames = list(map(lambda s: "USEquity(" + s + ").xlsm", factors))

    t0 = time.time()
    for f in filenames:
        wb = load_workbook(filename = f)
        sheet = wb.sheetnames[1]
        ws = wb[sheet]
        getAllTimeSeries(ws, sheet)
        print(time.time() - t0, "seconds wait time")
        wb.close()
    return

# returns a specific time series
def getTimeSeries(stock, worksheet, factor):
    stockList = wb["stock list"]
    i = findTickerIndex(stock, stockList)
    dates = []
    values = []
    row_range = 800
    letterIndex = convertIndexToLetter(i)
    for cell in worksheet['{}{}:{}{}'.format(letterIndex, 4, letterIndex, row_range)]:
        if cell[0].value is not None:
            dates.append(cell[0].value)
        else:
            break
    letterIndex = convertIndexToLetter(i+1)
    for cell in worksheet['{}{}:{}{}'.format(letterIndex, 4, letterIndex, row_range)]:
        if cell[0].value is not None:
            values.append(cell[0].value)
        else:
            break
    dates = np.array(dates, dtype = np.datetime64)
    data = pd.Series(values, index = dates, name = factor)
    return data


if __name__ == "__main__":
    wb = load_workbook(filename = "USEquity(Sector).xlsm", data_only=True)
    ws = wb["Sector"]
    getAllTickersBySector(ws)
